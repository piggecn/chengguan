# -*- coding: utf-8 -*-
"""社区城管日常巡查记录平台 — Flask 主程序。

移动端优先 · 账号角色权限（主管理员/办公室/中队长/队员） · SQLite 单文件 · 原图保留。
"""
import io
import os
import re
import secrets
import shutil
import sqlite3
import uuid
from datetime import datetime, timedelta
from functools import wraps
from pathlib import Path

from flask import (
    Flask, abort, flash, g, jsonify, redirect, render_template,
    request, send_file, send_from_directory, session, url_for,
)

import image_utils

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.environ.get("DATA_DIR", str(BASE_DIR / "data")))
UPLOADS_DIR = Path(os.environ.get("UPLOADS_DIR", str(BASE_DIR / "uploads")))
DATA_DIR.mkdir(parents=True, exist_ok=True)
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

ACCESS_PASSWORD = os.environ.get("ACCESS_PASSWORD", "").strip()

# ---------- 常量配置（增改一处全局生效） ----------
TEAMS = [
    "鄱阳镇社区一中队", "鄱阳镇社区二中队",
    "饶州街道社区一中队", "饶州街道社区二中队",
]


def get_units(kind=None):
    """单位列表（units 表，主管理员可维护）。kind: team/office/None(全部)。"""
    db = get_db()
    if kind:
        rows = db.execute(
            "SELECT * FROM units WHERE kind=? ORDER BY sort, id", (kind,)
        ).fetchall()
    else:
        rows = db.execute("SELECT * FROM units ORDER BY sort, id").fetchall()
    return [dict(r) for r in rows]


def team_units():
    return [u["name"] for u in get_units("team")]


def office_units():
    return [u["name"] for u in get_units("office")]


def all_unit_names():
    return [u["name"] for u in get_units()]


def unit_kind(name):
    row = get_db().execute(
        "SELECT kind FROM units WHERE name=?", (name,)
    ).fetchone()
    return row["kind"] if row else None


def town_units():
    """中队单位里出现过的乡镇（去重保序），用于台账按乡镇分组/筛选。"""
    towns = []
    for u in get_units("team"):
        if u["town"] and u["town"] not in towns:
            towns.append(u["town"])
    return towns


# 单位列表由主管理员在账号管理页维护（units 表）；TEAMS 仅作为首次初始化种子
CATEGORIES = [
    "违法搭建", "牛皮癣小广告", "乱堆放杂物", "电动车乱停放",
    "流动摊贩", "出店经营", "毁坏绿化", "占道经营",
    "破坏市政设施", "乱倒垃圾", "噪音扰民", "投诉纠纷", "其他",
]
PROGRESS_LABELS = {"investigating": "调查中", "filed": "已立案", "closed": "已办结"}
ROLE_LABELS = {"super": "主管理员", "office": "办公室管理员",
               "captain": "中队长", "vice-captain": "副中队长", "member": "队员"}
# 各角色可创建的账号角色
ADD_ROLES = {
    "super": [("captain", "中队长"), ("vice-captain", "副中队长"),
              ("member", "队员"), ("office", "办公室管理员")],
    "office": [("vice-captain", "副中队长"), ("member", "队员")],
    "captain": [("vice-captain", "副中队长"), ("member", "队员")],
}


def random_pin() -> str:
    return f"{secrets.randbelow(10000):04d}"


def get_setting(key, default=""):
    row = get_db().execute(
        "SELECT value FROM settings WHERE key=?", (key,)
    ).fetchone()
    return row["value"] if row else default


def set_setting(key, value):
    get_db().execute(
        "INSERT INTO settings(key, value) VALUES(?, ?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (key, value),
    )
    get_db().commit()


def base_url():
    """对外分享链接用的域名：主管理员在账号管理页配置，未配置则用当前访问域名。"""
    u = (get_setting("base_url") or "").strip().rstrip("/")
    if not u:
        u = request.host_url.rstrip("/")
    return u


def log_action(action, detail):
    """操作日志：记录是谁、做了什么。"""
    u = current_user()
    if not u:
        return
    get_db().execute(
        "INSERT INTO logs(user, action, detail, created_at) VALUES(?,?,?,?)",
        (u["name"], action, detail, now()),
    )
    get_db().commit()


app = Flask(__name__)


@app.context_processor
def inject_globals():
    """所有模板可用：user（当前用户或 None）、role_labels。"""
    return {"user": current_user(), "role_labels": ROLE_LABELS}


app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "chengguan-local-dev-secret")
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=180)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024  # 单次上传上限 100MB


# ---------- 数据库 ----------
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DATA_DIR / "records.db")
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(_exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    with sqlite3.connect(DATA_DIR / "records.db") as db:
        db.row_factory = sqlite3.Row
        db.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            unit TEXT NOT NULL,
            role TEXT NOT NULL,          -- super / office / captain / member
            title TEXT NOT NULL DEFAULT '',
            pin TEXT NOT NULL,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS units (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            kind TEXT NOT NULL DEFAULT 'team',  -- team=中队 / office=办公室
            town TEXT NOT NULL DEFAULT '',
            sort INTEGER NOT NULL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            team TEXT NOT NULL,
            community TEXT NOT NULL,
            category TEXT NOT NULL,
            description TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'pending',
            result TEXT DEFAULT '',
            deadline TEXT NOT NULL DEFAULT '',
            lead_dept TEXT NOT NULL DEFAULT '',
            assist_dept TEXT NOT NULL DEFAULT '',
            reporter TEXT NOT NULL,
            created_at TEXT NOT NULL,
            closed_at TEXT
        );
        CREATE TABLE IF NOT EXISTS images (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            record_id INTEGER NOT NULL,
            type TEXT NOT NULL,            -- before / after
            filepath TEXT NOT NULL,        -- 原图相对路径
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS communities (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            count INTEGER NOT NULL DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS cases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            team TEXT NOT NULL,
            case_no TEXT NOT NULL DEFAULT '',
            case_name TEXT NOT NULL,
            progress TEXT NOT NULL DEFAULT 'investigating',
            fine_amount REAL NOT NULL DEFAULT 0,
            reporter TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS complaints (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            team TEXT NOT NULL,
            community TEXT NOT NULL,
            phone TEXT DEFAULT '',
            content TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'pending',   -- pending / done
            reporter TEXT NOT NULL,
            handler TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            handled_at TEXT
        );
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user TEXT NOT NULL,
            action TEXT NOT NULL,
            detail TEXT NOT NULL,
            created_at TEXT NOT NULL
        );
        """)
        # 旧库迁移：records 缺列时补上
        cols = [r[1] for r in db.execute("PRAGMA table_info(records)").fetchall()]
        for col in ("deadline", "lead_dept", "assist_dept"):
            if col not in cols:
                db.execute(
                    "ALTER TABLE records ADD COLUMN %s TEXT NOT NULL DEFAULT ''" % col)
                print("[migrate] records 增加 %s 列" % col)
        ccols = [r[1] for r in db.execute("PRAGMA table_info(cases)").fetchall()]
        if "case_no" not in ccols:
            db.execute(
                "ALTER TABLE cases ADD COLUMN case_no TEXT NOT NULL DEFAULT ''")
            print("[migrate] cases 增加 case_no 列")
        ucols = [r[1] for r in db.execute("PRAGMA table_info(users)").fetchall()]
        if "title" not in ucols:
            db.execute(
                "ALTER TABLE users ADD COLUMN title TEXT NOT NULL DEFAULT ''")
            print("[migrate] users 增加 title 列")
        if db.execute("SELECT COUNT(*) c FROM units").fetchone()["c"] == 0:
            for i, t in enumerate(TEAMS):
                town = ""
                if t.startswith("鄱阳镇"):
                    town = "鄱阳镇"
                elif t.startswith("饶州街道"):
                    town = "饶州街道"
                db.execute(
                    "INSERT INTO units(name, kind, town, sort) VALUES(?,?,?,?)",
                    (t, "team", town, i))
            db.execute(
                "INSERT INTO units(name, kind, town, sort) "
                "VALUES('办公室','office','',99)")
            print("[init] 已初始化单位列表（4 中队 + 办公室）")
        # 主管理员不再挂「大队」，unit 置空走登录页单独入口
        db.execute("UPDATE users SET unit='' WHERE role='super'")
        # 投诉并入巡查（2026-09-02 拍板）：complaints 迁移为 records（分类「投诉纠纷」）
        def _thumb_rel(rel):
            stem, ext = rel.rsplit(".", 1)
            return f"{stem}_thumb.{ext}"

        for c in db.execute("SELECT * FROM complaints").fetchall():
            desc = (c["content"] or "").strip()
            if c["phone"]:
                desc = f"{desc}（来电：{c['phone']}）" if desc else f"来电：{c['phone']}"
            status = "closed" if c["status"] == "done" else "pending"
            result = ""
            if c["status"] == "done":
                result = "已处理"
                if c["handler"]:
                    result += f"（处理人：{c['handler']}）"
            cur = db.execute(
                "INSERT INTO records(team, community, category, description, "
                "status, result, deadline, reporter, created_at, closed_at) "
                "VALUES(?,?,?,?,?,?,?,?,?,?)",
                (c["team"], c["community"] or "", "投诉纠纷", desc, status,
                 result, "", c["reporter"] or "", c["created_at"],
                 c["handled_at"] if c["status"] == "done" else None),
            )
            rid = cur.lastrowid
            for im in db.execute(
                "SELECT * FROM images WHERE record_id=? AND type='complaint'",
                (c["id"],),
            ).fetchall():
                old_rel = im["filepath"]
                old_p = UPLOADS_DIR / old_rel
                new_rel = f"records/{rid}/before/{Path(old_rel).name}"
                new_p = UPLOADS_DIR / new_rel
                moved = False
                if old_p.exists():
                    new_p.parent.mkdir(parents=True, exist_ok=True)
                    shutil.move(str(old_p), str(new_p))
                    moved = True
                old_t = UPLOADS_DIR / _thumb_rel(old_rel)
                if old_t.exists():
                    new_t = UPLOADS_DIR / _thumb_rel(new_rel)
                    shutil.move(str(old_t), str(new_t))
                if moved:
                    db.execute(
                        "UPDATE images SET record_id=?, type='before', "
                        "filepath=? WHERE id=?",
                        (rid, new_rel, im["id"]))
                else:
                    db.execute("DELETE FROM images WHERE id=?", (im["id"],))
            db.execute("DELETE FROM complaints WHERE id=?", (c["id"],))
            print(f"[migrate] 投诉#{c['id']} → 巡查记录#{rid}（投诉纠纷）")
        cur = db.execute("SELECT COUNT(*) c FROM users")
        if cur.fetchone()["c"] == 0:
            db.execute(
                "INSERT INTO users(name, unit, role, pin, created_at) "
                "VALUES(?,?,?,?,?)",
                ("主管理员", "", "super", "0000", now()),
            )
            print("[init] 已创建唯一默认账号：主管理员（单位管理员），初始密码=0000，"
                  "其余账号由管理员在账号管理页自行添加")


def now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M")


init_db()


# ---------- 访问控制 ----------
@app.before_request
def gate_keeper():
    if ACCESS_PASSWORD and not session.get("gate_ok"):
        if request.endpoint not in ("gate", "static"):
            return redirect(url_for("gate"))


def current_user():
    uid = session.get("uid")
    if not uid:
        return None
    return get_db().execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()


def require_user(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not current_user():
            return redirect(url_for("personnel"))
        return view(*args, **kwargs)
    return wrapped


def scope_where(team_col="team", reporter_col="reporter"):
    """按角色返回数据可见范围 (where子句, 参数列表)。

    主管理员/办公室：全部；中队长/副中队长：本中队全部；队员：仅自己上报的。
    """
    u = current_user()
    if u["role"] in ("super", "office"):
        return "", []
    if u["role"] in ("captain", "vice-captain"):
        return f"{team_col} = ?", [u["unit"]]
    return f"{reporter_col} = ?", [u["name"]]


def can_view_record(r):
    u = current_user()
    if u["role"] in ("super", "office"):
        return True
    if u["role"] in ("captain", "vice-captain"):
        return r["team"] == u["unit"]
    return r["reporter"] == u["name"]


def can_view_case(c):
    u = current_user()
    if u["role"] in ("super", "office"):
        return True
    if u["role"] in ("captain", "vice-captain"):
        return c["team"] == u["unit"]
    return c["reporter"] == u["name"]


# ---------- 图片 ----------
def save_photos(files, subdir):
    base = UPLOADS_DIR / subdir
    base.mkdir(parents=True, exist_ok=True)
    saved = []
    for f in files:
        if not f or not f.filename:
            continue
        uid = uuid.uuid4().hex
        orig = f"{uid}.jpg"
        thumb = f"{uid}_thumb.jpg"
        try:
            image_utils.process(f, base / orig, base / thumb)
        except ValueError:
            continue
        saved.append((f"{subdir}/{orig}", f"{subdir}/{thumb}"))
    return saved


def thumb_of(filepath):
    stem, ext = filepath.rsplit(".", 1)
    return f"{stem}_thumb.{ext}"


def community_autocomplete(q, limit=8):
    db = get_db()
    if q:
        rows = db.execute(
            "SELECT name FROM communities WHERE name LIKE ? ORDER BY count DESC LIMIT ?",
            (f"%{q}%", limit),
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT name FROM communities ORDER BY count DESC LIMIT ?", (limit,)
        ).fetchall()
    return [r["name"] for r in rows]


def bump_community(name):
    db = get_db()
    db.execute(
        "INSERT INTO communities(name, count) VALUES(?, 1) "
        "ON CONFLICT(name) DO UPDATE SET count = count + 1",
        (name,),
    )
    db.commit()


# ---------- 站点密码门 ----------
@app.route("/gate", methods=["GET", "POST"])
def gate():
    if request.method == "POST":
        if request.form.get("password") == ACCESS_PASSWORD:
            session["gate_ok"] = True
            return redirect(url_for("index"))
        return render_template("gate.html", error="密码不对")
    return render_template("gate.html", error=None)


# ---------- 登录 ----------
@app.route("/personnel", methods=["GET", "POST"])
def personnel():
    if request.method == "POST":
        unit = (request.form.get("unit") or "").strip()
        name = (request.form.get("name") or "").strip()
        pin = (request.form.get("pin") or "").strip()
        if unit == "__super__":
            u = get_db().execute(
                "SELECT * FROM users WHERE role='super' AND name=? AND pin=?",
                (name, pin),
            ).fetchone()
        else:
            u = get_db().execute(
                "SELECT * FROM users WHERE unit=? AND name=? AND pin=?",
                (unit, name, pin),
            ).fetchone()
        if u:
            session.permanent = True
            session["uid"] = u["id"]
            return redirect(url_for("index"))
        return render_template("personnel.html", error="单位、姓名或密码不对",
                               units=get_units())
    return render_template("personnel.html", error=None, units=get_units())


@app.route("/api/users")
def api_users():
    unit = request.args.get("unit", "").strip()
    if unit == "__super__":
        rows = get_db().execute(
            "SELECT name FROM users WHERE role='super' ORDER BY id"
        ).fetchall()
        return jsonify([r["name"] for r in rows])
    if unit not in all_unit_names():
        return jsonify([])
    rows = get_db().execute(
        "SELECT name FROM users WHERE unit=? ORDER BY id", (unit,)
    ).fetchall()
    return jsonify([r["name"] for r in rows])


@app.route("/logout")
def logout():
    session.pop("uid", None)
    return redirect(url_for("personnel"))


# ---------- 账号管理（主管理员/办公室/中队长） ----------
def admin_access():
    """账号管理页面与操作权限校验，返回当前用户或 abort。"""
    u = current_user()
    if not u:
        return redirect(url_for("personnel"))
    if u["role"] not in ("super", "office", "captain"):
        abort(403)
    return u


@app.route("/admin")
@require_user
def admin():
    u = admin_access()
    db = get_db()
    if u["role"] == "captain":
        users = db.execute(
            "SELECT * FROM users WHERE unit=? ORDER BY id", (u["unit"],)
        ).fetchall()
        add_units = [u["unit"]]
    elif u["role"] == "office":
        users = db.execute(
            "SELECT * FROM users ORDER BY unit, id"
        ).fetchall()
        add_units = team_units()
    else:  # super
        users = db.execute(
            "SELECT * FROM users ORDER BY unit, id"
        ).fetchall()
        add_units = all_unit_names()

    # 密码可见范围：主管理员看全部；办公室看不到主管理员的；中队长只看本队队员(副中队长/队员)
    share_base = base_url()
    items = []
    for row in users:
        row = dict(row)
        if u["role"] == "super":
            row["show_pin"] = True
        elif u["role"] == "office":
            row["show_pin"] = row["role"] != "super"
        else:  # captain
            row["show_pin"] = row["role"] in ("vice-captain", "member")
        row["can_reset"] = u["role"] == "super"  # 重置密码仅主管理员
        if row["show_pin"]:
            row["share_url"] = (
                share_base + url_for("personnel",
                                     unit=row["unit"], name=row["name"],
                                     pin=row["pin"])
            )
        items.append(row)

    logs = []
    if u["role"] == "super":
        logs = get_db().execute(
            "SELECT * FROM logs ORDER BY id DESC LIMIT 100"
        ).fetchall()

    return render_template("admin.html", users=items, teams=team_units(),
                           role_labels=ROLE_LABELS, user=u,
                           add_units=add_units, add_unit_kinds={
                               n: unit_kind(n) for n in add_units},
                           unit_list=get_units(),
                           base_url=get_setting("base_url", ""),
                           logs=logs,
                           team_roles=[r for r in ADD_ROLES[u["role"]] if r[0] != "office"],
                           office_roles=[r for r in ADD_ROLES[u["role"]] if r[0] == "office"])


@app.route("/admin/unit/add", methods=["POST"])
@require_user
def admin_unit_add():
    u = admin_access()
    if u["role"] != "super":
        flash("只有主管理员可以管理单位", "error")
        return redirect(url_for("admin"))
    name = (request.form.get("name") or "").strip()
    kind = request.form.get("kind", "team").strip()
    town = (request.form.get("town") or "").strip()
    if not name:
        flash("单位名称不能为空", "error")
        return redirect(url_for("admin"))
    if kind not in ("team", "office"):
        kind = "team"
    if get_db().execute("SELECT 1 FROM units WHERE name=?", (name,)).fetchone():
        flash("已存在同名单位", "error")
        return redirect(url_for("admin"))
    sort = len(get_units()) + 1
    get_db().execute(
        "INSERT INTO units(name, kind, town, sort) VALUES(?,?,?,?)",
        (name, kind, town, sort),
    )
    get_db().commit()
    log_action("新增单位", f"{name}（{'中队' if kind == 'team' else '办公室'}）")
    flash(f"单位「{name}」已添加", "ok")
    return redirect(url_for("admin"))


@app.route("/admin/unit/rename", methods=["POST"])
@require_user
def admin_unit_rename():
    u = admin_access()
    if u["role"] != "super":
        abort(403)
    uid = request.form.get("uid", "")
    new_name = (request.form.get("name") or "").strip()
    row = get_db().execute(
        "SELECT * FROM units WHERE id=?", (uid,)
    ).fetchone() if uid.isdigit() else None
    if not row or not new_name:
        flash("参数不对", "error")
        return redirect(url_for("admin"))
    if new_name != row["name"] and get_db().execute(
            "SELECT 1 FROM units WHERE name=?", (new_name,)).fetchone():
        flash("已存在同名单位", "error")
        return redirect(url_for("admin"))
    db = get_db()
    # 历史记录保留旧名（数据不动），仅更新单位表和该单位下账号的单位名
    db.execute("UPDATE users SET unit=? WHERE unit=?", (new_name, row["name"]))
    db.execute("UPDATE units SET name=?, town=? WHERE id=?",
               (new_name, (request.form.get("town") or row["town"]).strip(),
                uid))
    db.commit()
    log_action("单位改名", f"{row['name']} → {new_name}")
    flash("单位已更新", "ok")
    return redirect(url_for("admin"))


@app.route("/admin/unit/delete", methods=["POST"])
@require_user
def admin_unit_delete():
    u = admin_access()
    if u["role"] != "super":
        abort(403)
    uid = request.form.get("uid", "")
    row = get_db().execute(
        "SELECT * FROM units WHERE id=?", (uid,)
    ).fetchone() if uid.isdigit() else None
    if not row:
        flash("参数不对", "error")
        return redirect(url_for("admin"))
    cnt = get_db().execute(
        "SELECT COUNT(*) c FROM users WHERE unit=?", (row["name"],)
    ).fetchone()["c"]
    if cnt:
        flash(f"「{row['name']}」下还有 {cnt} 个账号，先删除账号才能删单位", "error")
        return redirect(url_for("admin"))
    get_db().execute("DELETE FROM units WHERE id=?", (uid,))
    get_db().commit()
    log_action("删除单位", row["name"])
    flash(f"单位「{row['name']}」已删除", "ok")
    return redirect(url_for("admin"))


@app.route("/admin/settings", methods=["POST"])
@require_user
def admin_settings():
    u = admin_access()
    if u["role"] != "super":
        flash("只有主管理员可以配置域名", "error")
        return redirect(url_for("admin"))
    val = (request.form.get("base_url") or "").strip().rstrip("/")
    if val and not val.startswith(("http://", "https://")):
        flash("域名要以 http:// 或 https:// 开头，例如 http://192.168.50.65:8755", "error")
        return redirect(url_for("admin"))
    set_setting("base_url", val)
    flash("分享域名已保存，分享链接将使用：" + (val or "当前访问地址"), "ok")
    return redirect(url_for("admin"))


@app.route("/admin/add", methods=["POST"])
@require_user
def admin_add():
    u = admin_access()
    name = (request.form.get("name") or "").strip()
    unit = request.form.get("unit", "")
    role = request.form.get("role", "")
    if not name:
        flash("姓名不能为空", "error")
        return redirect(url_for("admin"))
    if u["role"] == "captain":
        if unit and unit != u["unit"]:
            flash("只能给自己中队添加人员", "error")
            return redirect(url_for("admin"))
        unit = u["unit"]
    if u["role"] in ("office", "captain") and role not in ("vice-captain", "member"):
        flash("只能添加副中队长和队员", "error")
        return redirect(url_for("admin"))
    title = (request.form.get("title") or "").strip()
    kind = unit_kind(unit)
    if u["role"] == "super":
        if kind == "office":
            valid = role == "office"
        elif kind == "team":
            valid = role in ("captain", "vice-captain", "member")
        else:
            valid = False
    elif u["role"] == "office":
        valid = kind == "team" and role in ("vice-captain", "member")
    else:
        valid = unit == u["unit"] and role in ("vice-captain", "member")
    if not valid:
        flash("单位与角色组合不合法", "error")
        return redirect(url_for("admin"))
    if role == "office" and title not in ("办公室主任", "办公室科员"):
        title = "办公室主任"
    if role != "office":
        title = ""
    if get_db().execute(
            "SELECT 1 FROM users WHERE unit=? AND name=?", (unit, name)
    ).fetchone():
        flash(f"{unit} 已存在同名人员", "error")
        return redirect(url_for("admin"))
    pin = random_pin()
    get_db().execute(
        "INSERT INTO users(name, unit, role, title, pin, created_at) "
        "VALUES(?,?,?,?,?,?)",
        (name, unit, role, title, pin, now()),
    )
    get_db().commit()
    log_action("创建账号", f"{unit} · {name}（{ROLE_LABELS[role]}）")
    flash(f"已创建 {unit}·{name}（{ROLE_LABELS[role]}），初始密码 {pin}，请转交本人", "ok")
    return redirect(url_for("admin"))


@app.route("/admin/reset/<int:uid>", methods=["POST"])
@require_user
def admin_reset(uid):
    u = admin_access()
    if u["role"] != "super":
        flash("只有主管理员可以重置密码", "error")
        return redirect(url_for("admin"))
    target = get_db().execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not target:
        abort(404)
    pin = random_pin()
    get_db().execute("UPDATE users SET pin=? WHERE id=?", (pin, uid))
    get_db().commit()
    log_action("重置密码", f"{target['name']}（{target['unit']}）")
    flash(f"已重置 {target['name']} 密码，新密码 {pin}，请转交本人", "ok")
    return redirect(url_for("admin"))


# ---------- 修改自己的密码（所有人） ----------
@app.route("/password", methods=["GET", "POST"])
@require_user
def password():
    u = current_user()
    if request.method == "POST":
        cur = (request.form.get("current") or "").strip()
        new = (request.form.get("new") or "").strip()
        confirm = (request.form.get("confirm") or "").strip()
        if u["pin"] != cur:
            return render_template("password.html", error="当前密码不对")
        if len(new) != 4 or not new.isdigit():
            return render_template("password.html", error="新密码必须是 4 位数字")
        if new != confirm:
            return render_template("password.html", error="两次输入的新密码不一致")
        get_db().execute("UPDATE users SET pin=? WHERE id=?", (new, u["id"]))
        get_db().commit()
        log_action("修改密码", "修改自己的 4 位密码")
        flash("密码修改成功", "ok")
        return redirect(url_for("index"))
    return render_template("password.html", error=None)


# ---------- 巡查记录 ----------
@app.route("/")
@require_user
def index():
    u = current_user()
    team = request.args.get("team", "")
    category = request.args.get("category", "")
    reporter = request.args.get("reporter", "")
    status = request.args.get("status", "")
    month = request.args.get("month", "")
    community = request.args.get("community", "")
    q = (request.args.get("q") or "").strip()

    where, params = scope_where()
    if where and team:
        where += " AND team = ?"
        params.append(team)
    elif where:
        pass  # 中队长/队员的 team 已由 scope 限定
    else:
        if team:
            where = "team = ?"
            params = [team]
    if category:
        where = (where + " AND " if where else "") + "category = ?"
        params.append(category)
    if reporter:
        where = (where + " AND " if where else "") + "reporter = ?"
        params.append(reporter)
    if status in ("pending", "closed"):
        where = (where + " AND " if where else "") + "status = ?"
        params.append(status)
    if month:
        where = (where + " AND " if where else "") + "created_at LIKE ?"
        params.append(month + "%")
    if community:
        where = (where + " AND " if where else "") + "community = ?"
        params.append(community)
    if q:
        where = (where + " AND " if where else "") + "(community LIKE ? OR description LIKE ?)"
        params += [f"%{q}%", f"%{q}%"]

    sql = "SELECT * FROM records"
    if where:
        sql += " WHERE " + where
    sql += " ORDER BY id DESC"
    records = [dict(r) for r in get_db().execute(sql, params).fetchall()]
    today = datetime.now().date()
    for r in records:
        img = get_db().execute(
            "SELECT filepath FROM images WHERE record_id=? AND type='before' "
            "ORDER BY id LIMIT 1", (r["id"],)
        ).fetchone()
        r["thumb"] = thumb_of(img["filepath"]) if img else None
        if r["status"] == "pending":
            try:
                d = datetime.strptime(r["created_at"][:10], "%Y-%m-%d").date()
                r["days_pending"] = (today - d).days
            except ValueError:
                r["days_pending"] = 0
        else:
            r["days_pending"] = None

    # 统计也按权限范围
    this_month = datetime.now().strftime("%Y-%m")
    stat_where = where
    stat_params = list(params)
    def count_where(extra=""):
        w = stat_where
        if extra:
            w = (w + " AND " if w else "") + extra
        sql2 = "SELECT COUNT(*) c FROM records"
        if w:
            sql2 += " WHERE " + w
        return get_db().execute(sql2, stat_params).fetchone()["c"]

    stats = {
        "month_new": count_where("created_at LIKE ?") if False else
            get_db().execute(
                "SELECT COUNT(*) c FROM records WHERE " +
                ((stat_where + " AND ") if stat_where else "") +
                "created_at LIKE ?", stat_params + [this_month + "%"],
            ).fetchone()["c"],
        "pending": get_db().execute(
            "SELECT COUNT(*) c FROM records WHERE " +
            ((stat_where + " AND ") if stat_where else "") + "status='pending'",
            stat_params,
        ).fetchone()["c"],
        "closed": get_db().execute(
            "SELECT COUNT(*) c FROM records WHERE " +
            ((stat_where + " AND ") if stat_where else "") + "status='closed'",
            stat_params,
        ).fetchone()["c"],
    }

    # 筛选下拉选项
    can_all = u["role"] in ("super", "office")
    team_options = team_units() if can_all else []
    if u["role"] in ("captain", "vice-captain"):
        reporter_rows = get_db().execute(
            "SELECT name FROM users WHERE unit=? ORDER BY id", (u["unit"],)
        ).fetchall()
    elif can_all:
        reporter_rows = get_db().execute(
            "SELECT name FROM users ORDER BY unit, id"
        ).fetchall()
    else:
        reporter_rows = []

    return render_template(
        "index.html", records=records, stats=stats, user=u,
        categories=CATEGORIES, role_labels=ROLE_LABELS,
        team_options=team_options, reporter_options=[r["name"] for r in reporter_rows],
        can_all=can_all,
        sel={"team": team, "category": category, "reporter": reporter,
             "status": status, "month": month, "community": community, "q": q},
        this_month=this_month,
    )


@app.route("/create", methods=["GET", "POST"])
@require_user
def create():
    u = current_user()
    if request.method == "POST":
        if u["role"] in ("captain", "vice-captain", "member"):
            team = u["unit"]
        else:
            team = (request.form.get("team") or "").strip()
        community = (request.form.get("community") or "").strip()
        category = (request.form.get("category") or "").strip() or "其他"
        description = (request.form.get("description") or "").strip()
        deadline = (request.form.get("deadline") or "").strip()
        lead_dept = (request.form.get("lead_dept") or "").strip()
        assist_dept = (request.form.get("assist_dept") or "").strip()
        if team not in team_units():
            return render_template("create.html", error="请选择所属中队",
                                   teams=team_units(), categories=CATEGORIES,
                                   user=u), 400
        # 小区、分类、描述都可以留空（分类缺省记“其他”），传了之后可在详情页编辑补全
        db = get_db()
        cur = db.execute(
            "INSERT INTO records(team, community, category, description, "
            "status, deadline, lead_dept, assist_dept, reporter, created_at) "
            "VALUES(?,?,?,?,?,?,?,?,?,?)",
            (team, community, category, description, "pending", deadline,
             lead_dept, assist_dept, u["name"], now()),
        )
        rid = cur.lastrowid
        saved = save_photos(request.files.getlist("photos"),
                            f"records/{rid}/before")
        db.executemany(
            "INSERT INTO images(record_id, type, filepath, created_at) "
            "VALUES(?,?,?,?)",
            [(rid, "before", p[0], now()) for p in saved],
        )
        db.commit()
        log_action("新增巡查记录",
                   f"{team} · {community or '未填小区'} · {category}")
        bump_community(community)
        return redirect(url_for("detail", rid=rid))
    return render_template(
        "create.html", error=None, teams=team_units(), categories=CATEGORIES,
        user=u, old=None,
        lead_default=get_setting("ledger_lead_dept", "县城市管理综合行政执法大队"),
        assist_default=get_setting("ledger_assist_dept", "社区、物业"))


@app.route("/record/<int:rid>")
@require_user
def detail(rid):
    r = get_db().execute("SELECT * FROM records WHERE id=?", (rid,)).fetchone()
    if not r:
        abort(404)
    if not can_view_record(r):
        abort(403)
    images = get_db().execute(
        "SELECT * FROM images WHERE record_id=? ORDER BY id", (rid,)
    ).fetchall()
    before = [dict(i) for i in images if i["type"] == "before"]
    after = [dict(i) for i in images if i["type"] == "after"]
    return render_template("detail.html", r=r, before=before, after=after,
                           progress_labels=PROGRESS_LABELS,
                           user=current_user())


@app.route("/record/<int:rid>/edit", methods=["GET", "POST"])
@require_user
def edit_record(rid):
    u = current_user()
    r = get_db().execute("SELECT * FROM records WHERE id=?", (rid,)).fetchone()
    if not r:
        abort(404)
    if not can_view_record(r):
        abort(403)
    if request.method == "POST":
        team = (request.form.get("team") or "").strip()
        community = (request.form.get("community") or "").strip()
        category = (request.form.get("category") or "").strip() or "其他"
        description = (request.form.get("description") or "").strip()
        deadline = (request.form.get("deadline") or "").strip()
        lead_dept = (request.form.get("lead_dept") or "").strip()
        assist_dept = (request.form.get("assist_dept") or "").strip()
        if team not in team_units():
            return render_template("edit.html", r=r, teams=team_units(),
                                   categories=CATEGORIES, user=u,
                                   error="请选择所属中队"), 400
        get_db().execute(
            "UPDATE records SET team=?, community=?, category=?, description=?, "
            "deadline=?, lead_dept=?, assist_dept=? WHERE id=?",
            (team, community, category, description, deadline, lead_dept,
             assist_dept, rid),
        )
        get_db().commit()
        log_action("编辑巡查记录", f"记录#{rid} {community or '未填小区'} · {category}")
        bump_community(community)
        return redirect(url_for("detail", rid=rid))
    return render_template(
        "edit.html", r=r, teams=team_units(), categories=CATEGORIES,
        user=u, error=None,
        lead_default=get_setting("ledger_lead_dept", "县城市管理综合行政执法大队"),
        assist_default=get_setting("ledger_assist_dept", "社区、物业"))


@app.route("/record/<int:rid>/delete", methods=["POST"])
@require_user
def delete_record(rid):
    r = get_db().execute("SELECT * FROM records WHERE id=?", (rid,)).fetchone()
    if not r:
        abort(404)
    if not can_view_record(r):
        abort(403)
    images = get_db().execute(
        "SELECT filepath FROM images WHERE record_id=?", (rid,)
    ).fetchall()
    db = get_db()
    db.execute("DELETE FROM images WHERE record_id=?", (rid,))
    db.execute("DELETE FROM records WHERE id=?", (rid,))
    db.commit()
    log_action("删除巡查记录",
               f"记录#{rid} {r['community'] or '未填小区'} · {r['category']}")
    for img in images:
        p = UPLOADS_DIR / img["filepath"]
        try:
            if p.exists():
                p.unlink()
            t = UPLOADS_DIR / thumb_of(img["filepath"])
            if t.exists():
                t.unlink()
        except OSError:
            pass
    flash("记录已删除", "ok")
    return redirect(url_for("index"))


@app.route("/record/<int:rid>/close", methods=["GET", "POST"])
@require_user
def close(rid):
    r = get_db().execute("SELECT * FROM records WHERE id=?", (rid,)).fetchone()
    if not r:
        abort(404)
    if not can_view_record(r):
        abort(403)
    if r["status"] == "closed":
        return redirect(url_for("detail", rid=rid))
    if request.method == "POST":
        result = (request.form.get("result") or "").strip()
        if not result:
            return render_template("close.html", r=r, error="处理结果要填")
        deadline = (request.form.get("deadline") or "").strip() or r["deadline"]
        db = get_db()
        saved = save_photos(request.files.getlist("photos"),
                            f"records/{rid}/after")
        db.executemany(
            "INSERT INTO images(record_id, type, filepath, created_at) "
            "VALUES(?,?,?,?)",
            [(rid, "after", p[0], now()) for p in saved],
        )
        db.execute(
            "UPDATE records SET status='closed', result=?, closed_at=?, "
            "deadline=? WHERE id=?", (result, now(), deadline, rid),
        )
        db.commit()
        log_action("整改销号", f"记录#{rid} {r['community'] or '未填小区'} · {r['category']}")
        return redirect(url_for("detail", rid=rid))
    return render_template("close.html", r=r, error=None)


# ---------- 案件（并入巡查主流程） ----------
@app.route("/cases", methods=["GET", "POST"])
@require_user
def cases():
    u = current_user()
    db = get_db()
    if request.method == "POST":
        if u["role"] in ("captain", "vice-captain", "member"):
            team = u["unit"]
        else:
            team = (request.form.get("team") or "").strip()
        case_no = (request.form.get("case_no") or "").strip()
        case_name = (request.form.get("case_name") or "").strip()
        progress = request.form.get("progress", "investigating")
        fine = request.form.get("fine_amount", "0").strip() or "0"
        if team not in team_units() or not case_name:
            return render_template(
                "cases.html", error="中队和案件名称要填", rows=[],
                teams=team_units(), progress_labels=PROGRESS_LABELS, user=u,
            ), 400
        db.execute(
            "INSERT INTO cases(team, case_no, case_name, progress, "
            "fine_amount, reporter, created_at, updated_at) "
            "VALUES(?,?,?,?,?,?,?,?)",
            (team, case_no, case_name, progress, float(fine), u["name"],
             now(), now()),
        )
        db.commit()
        log_action("登记案件", f"{team} · {case_name}")
        return redirect(url_for("cases"))

    month = request.args.get("month", "")
    team_q = request.args.get("team", "")
    progress_q = request.args.get("progress", "")

    where, params = scope_where()
    if team_q:
        where = (where + " AND " if where else "") + "team = ?"
        params.append(team_q)
    if progress_q in PROGRESS_LABELS:
        where = (where + " AND " if where else "") + "progress = ?"
        params.append(progress_q)
    if month:
        where = (where + " AND " if where else "") + "created_at LIKE ?"
        params.append(month + "%")
    sql = "SELECT * FROM cases"
    if where:
        sql += " WHERE " + where
    sql += " ORDER BY id DESC"
    rows = db.execute(sql, params).fetchall()
    return render_template("cases.html", error=None, rows=rows,
                           teams=team_units(), progress_labels=PROGRESS_LABELS,
                           user=u, can_all=u["role"] in ("super", "office"),
                           sel={"team": team_q, "progress": progress_q,
                                "month": month})


@app.route("/case/<int:cid>/update", methods=["POST"])
@require_user
def case_update(cid):
    c = get_db().execute("SELECT * FROM cases WHERE id=?", (cid,)).fetchone()
    if not c or not can_view_case(c):
        abort(404)
    progress = request.form.get("progress", "investigating")
    fine = (request.form.get("fine_amount") or "").strip() or "0"
    db = get_db()
    db.execute(
        "UPDATE cases SET progress=?, fine_amount=?, updated_at=? WHERE id=?",
        (progress, float(fine), now(), cid),
    )
    db.commit()
    log_action("更新案件", f"案件#{cid} 进度={progress} 罚款={fine}")
    return redirect(url_for("cases"))


@app.route("/case/<int:cid>/edit", methods=["GET", "POST"])
@require_user
def case_edit(cid):
    u = current_user()
    c = get_db().execute("SELECT * FROM cases WHERE id=?", (cid,)).fetchone()
    if not c or not can_view_case(c):
        abort(404)
    if request.method == "POST":
        if u["role"] in ("super", "office"):
            team = (request.form.get("team") or "").strip()
        else:
            team = c["team"]
        case_no = (request.form.get("case_no") or "").strip()
        case_name = (request.form.get("case_name") or "").strip()
        fine = (request.form.get("fine_amount") or "").strip() or "0"
        if team not in team_units() or not case_name:
            return render_template("case_edit.html", c=c, teams=team_units(),
                                   user=u, error="中队和案件名称要填"), 400
        db = get_db()
        db.execute(
            "UPDATE cases SET team=?, case_no=?, case_name=?, fine_amount=?, "
            "updated_at=? WHERE id=?",
            (team, case_no, case_name, float(fine), now(), cid),
        )
        db.commit()
        log_action("编辑案件", f"案件#{cid} {case_name}")
        return redirect(url_for("cases"))
    return render_template("case_edit.html", c=c, teams=team_units(), user=u, error=None)


@app.route("/case/<int:cid>/delete", methods=["POST"])
@require_user
def case_delete(cid):
    c = get_db().execute("SELECT * FROM cases WHERE id=?", (cid,)).fetchone()
    if not c or not can_view_case(c):
        abort(404)
    get_db().execute("DELETE FROM cases WHERE id=?", (cid,))
    get_db().commit()
    log_action("删除案件", f"案件#{cid} {c['case_name']}")
    flash("案件已删除", "ok")
    return redirect(url_for("cases"))


# ---------- 统计 ----------
@app.route("/stats")
@require_user
def stats():
    month = request.args.get("month") or datetime.now().strftime("%Y-%m")
    prefix = month + "%"
    where, params = scope_where()
    rw = (where + " AND ") if where else ""
    db = get_db()

    total = db.execute(
        f"SELECT COUNT(*) c FROM records WHERE {rw}created_at LIKE ?",
        params + [prefix],
    ).fetchone()["c"]
    closed = db.execute(
        f"SELECT COUNT(*) c FROM records WHERE {rw}created_at LIKE ? "
        "AND status='closed'", params + [prefix],
    ).fetchone()["c"]

    by_team = db.execute(
        f"SELECT team, COUNT(*) c, SUM(status='closed') closed FROM records "
        f"WHERE {rw}created_at LIKE ? GROUP BY team", params + [prefix],
    ).fetchall()
    by_category = db.execute(
        f"SELECT category, COUNT(*) c FROM records WHERE {rw}created_at LIKE ? "
        "GROUP BY category ORDER BY c DESC", params + [prefix],
    ).fetchall()
    by_person = db.execute(
        f"SELECT reporter, COUNT(*) c, SUM(status='closed') closed FROM records "
        f"WHERE {rw}created_at LIKE ? GROUP BY reporter ORDER BY c DESC",
        params + [prefix],
    ).fetchall()
    by_community = db.execute(
        f"SELECT community, COUNT(*) c FROM records WHERE {rw}created_at LIKE ? "
        "GROUP BY community ORDER BY c DESC LIMIT 10", params + [prefix],
    ).fetchall()

    case_total = db.execute(
        f"SELECT COUNT(*) c FROM cases WHERE {rw}created_at LIKE ?",
        params + [prefix],
    ).fetchone()["c"]
    case_fine = db.execute(
        f"SELECT COALESCE(SUM(fine_amount),0) s FROM cases WHERE {rw}created_at LIKE ?",
        params + [prefix],
    ).fetchone()["s"]
    case_by_progress = db.execute(
        f"SELECT progress, COUNT(*) c FROM cases WHERE {rw}created_at LIKE ? "
        "GROUP BY progress", params + [prefix],
    ).fetchall()
    case_by_team = db.execute(
        f"SELECT team, COUNT(*) c FROM cases WHERE {rw}created_at LIKE ? "
        "GROUP BY team", params + [prefix],
    ).fetchall()

    return render_template(
        "stats.html", month=month,
        total=total, closed=closed,
        rate=round(closed / total * 100, 1) if total else 0,
        by_team=by_team, by_category=by_category,
        by_person=by_person, by_community=by_community,
        case_total=case_total, case_fine=case_fine,
        case_by_progress=case_by_progress, case_by_team=case_by_team,
        progress_labels=PROGRESS_LABELS, user=current_user(),
    )


# ---------- 导出（Excel 文字 + 原图打包 ZIP） ----------
@app.route("/export")
@require_user
def export_data():
    from openpyxl import Workbook
    import zipfile as _zip

    team = request.args.get("team", "")
    category = request.args.get("category", "")
    reporter = request.args.get("reporter", "")
    status = request.args.get("status", "")
    month = request.args.get("month", "")
    community = request.args.get("community", "")
    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()

    where, params = scope_where()
    if team:
        where = (where + " AND " if where else "") + "team = ?"
        params.append(team)
    if category:
        where = (where + " AND " if where else "") + "category = ?"
        params.append(category)
    if reporter:
        where = (where + " AND " if where else "") + "reporter = ?"
        params.append(reporter)
    if status in ("pending", "closed"):
        where = (where + " AND " if where else "") + "status = ?"
        params.append(status)
    if month:
        where = (where + " AND " if where else "") + "created_at LIKE ?"
        params.append(month + "%")
    if community:
        where = (where + " AND " if where else "") + "community = ?"
        params.append(community)
    if start:
        where = (where + " AND " if where else "") + "created_at >= ?"
        params.append(start + " 00:00")
    if end:
        where = (where + " AND " if where else "") + "created_at <= ?"
        params.append(end + " 23:59")

    sql = "SELECT * FROM records"
    if where:
        sql += " WHERE " + where
    sql += " ORDER BY id"
    records = [dict(r) for r in get_db().execute(sql, params).fetchall()]

    # 按问题分类分组排序（按 CATEGORIES 既定顺序，组内按小区+时间）
    order = {c: i for i, c in enumerate(CATEGORIES)}
    records.sort(key=lambda r: (order.get(r["category"], 99),
                                r["community"], r["id"]))

    headers = ["序号", "所属中队", "小区", "问题分类", "问题描述", "状态",
               "上报人", "上报时间", "销号时间", "处理结果", "照片张数"]
    widths = [6, 20, 18, 14, 44, 8, 10, 17, 17, 30, 8]

    def photo_count(rid):
        return get_db().execute(
            "SELECT COUNT(*) c FROM images WHERE record_id=?", (rid,)
        ).fetchone()["c"]

    def row_of(i, r):
        return [i, r["team"], r["community"], r["category"], r["description"],
                "已办结" if r["status"] == "closed" else "待处理",
                r["reporter"], r["created_at"], r["closed_at"] or "",
                r["result"] or "", photo_count(r["id"])]

    def fill_sheet(ws, rows):
        ws.append(headers)
        for row in rows:
            ws.append(row)
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
        ws.freeze_panes = "A2"

    wb = Workbook()
    # 汇总表
    fill_sheet(wb.active, [row_of(i, r) for i, r in enumerate(records, 1)])
    wb.active.title = "汇总"
    # 每个问题分类一个工作表（只建非空分类）
    by_cat = {}
    for r in records:
        by_cat.setdefault(r["category"], []).append(r)
    for cat in CATEGORIES:
        if cat not in by_cat:
            continue
        ws = wb.create_sheet(title=cat)
        fill_sheet(ws, [row_of(i, r) for i, r in enumerate(by_cat[cat], 1)])

    buf = io.BytesIO()
    with _zip.ZipFile(buf, "w", _zip.ZIP_DEFLATED) as z:
        xbuf = io.BytesIO()
        wb.save(xbuf)
        label = f"巡查记录_{month or '全部'}.xlsx"
        z.writestr(label, xbuf.getvalue())
        for i, r in enumerate(records, 1):
            imgs = get_db().execute(
                "SELECT * FROM images WHERE record_id=? ORDER BY id", (r["id"],)
            ).fetchall()
            if not imgs:
                continue
            for j, img in enumerate(imgs, 1):
                p = UPLOADS_DIR / img["filepath"]
                if not p.exists():
                    continue
                typ = "现场" if img["type"] == "before" else "整改"
                z.write(p, f"照片/{r['category']}/{i}_{r['community']}_{typ}{j}.jpg")
    buf.seek(0)
    range_label = month or (f"{start}_to_{end}" if start or end else "全部")
    fname = f"巡查导出_{range_label}.zip".replace("/", "-")
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/zip")


# ---------- 导出筛选参数（通用） ----------
def _export_filters():
    """解析导出通用筛选参数，返回 (where, params, start, end, month)。"""
    team = request.args.get("team", "")
    town = request.args.get("town", "")
    category = request.args.get("category", "")
    reporter = request.args.get("reporter", "")
    status = request.args.get("status", "")
    month = request.args.get("month", "")
    community = request.args.get("community", "")
    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()
    where, params = scope_where()
    if town:
        t_names = [u["name"] for u in get_units("team") if u["town"] == town]
        if t_names:
            where = (where + " AND " if where else "") +                 "team IN (%s)" % ",".join("?" * len(t_names))
            params += t_names
    for cond, p in [
        (team, ("team = ?", team)),
        (category, ("category = ?", category)),
        (reporter, ("reporter = ?", reporter)),
        (status if status in ("pending", "closed") else "", ("status = ?", status)),
    ]:
        if cond:
            where = (where + " AND " if where else "") + p[0]
            params.append(p[1])
    if month:
        where = (where + " AND " if where else "") + "created_at LIKE ?"
        params.append(month + "%")
    if community:
        where = (where + " AND " if where else "") + "community = ?"
        params.append(community)
    if start:
        where = (where + " AND " if where else "") + "created_at >= ?"
        params.append(start + " 00:00")
    if end:
        where = (where + " AND " if where else "") + "created_at <= ?"
        params.append(end + " 23:59")
    return where, params, start, end, month


# ---------- 小区摸排台账导出（预览页 + Excel） ----------
LEDGER_DEPS = [
    ("ledger_report_dept", "问题上报部门", "县城管局"),
    ("ledger_lead_dept", "牵头部门", "县城市管理综合行政执法大队"),
    ("ledger_assist_dept", "配合部门", "社区、物业"),
]


def _ledger_groups_from(where, params, unit=None, unit_town=False):
    """按给定 SQL 条件取巡查记录并按小区分组；居民投诉并入同小区表（类目「居民投诉」共用序号）。"""
    sql = "SELECT * FROM records"
    if where:
        sql += " WHERE " + where
    records = [dict(r) for r in get_db().execute(sql, params).fetchall()]

    order = {c: i for i, c in enumerate(CATEGORIES)}
    by_comm = {}
    for r in records:
        by_comm.setdefault(r["community"] or "未填小区", []).append(r)
    groups = []
    for comm in sorted(by_comm):
        recs = sorted(by_comm.get(comm, []),
                      key=lambda r: (order.get(r["category"], 99), r["id"]))
        rows = []
        num, last_cat = 0, None
        blocks = []  # 每组照片一块：共用该记录分类的序号
        for r in recs:
            if r["category"] != last_cat:
                num += 1
                last_cat = r["category"]
            r["_status_text"] = "已整改" if r["status"] == "closed" else "未整改"
            r["_remark"] = ""
            rows.append((num, r))
            b, a = [], []
            for im in get_db().execute(
                "SELECT * FROM images WHERE record_id=? ORDER BY id",
                (r["id"],),
            ).fetchall():
                (b if im["type"] == "before" else a).append(im["filepath"])
            if b or a:
                blocks.append({"num": num, "before": b, "after": a,
                               "before_thumbs": [thumb_of(p) for p in b],
                               "after_thumbs": [thumb_of(p) for p in a]})
        groups.append({
            "community": comm, "rows": rows,
            "pad": max(0, 9 - len(rows)),  # 预览/表格固定 9 行序号空间
            "blocks": blocks,
        })
    return groups


def _ledger_groups():
    """摸排台账数据：按小区分组，每小区返回序号共用行 + 整改前/后照片路径。"""
    where, params, start, end, month = _export_filters()
    return _ledger_groups_from(where, params), start, end


@app.route("/export/ledger")
@require_user
def export_ledger():
    groups, start, end = _ledger_groups()
    deps = {k: get_setting(k, d) or d for k, _l, d in LEDGER_DEPS}
    # 权限范围内出现过的小区，供「打印特定小区」下拉
    where, params = scope_where()
    sql = "SELECT DISTINCT community FROM records WHERE community != ''"
    if where:
        sql += " AND " + where
    cnames = {r["community"] for r in get_db().execute(sql, params).fetchall()}
    return render_template(
        "ledger.html", groups=groups, deps=deps,
        start=start, end=end,
        sel={"town": request.args.get("town", ""),
             "team": request.args.get("team", ""),
             "group": request.args.get("group", ""),
             "community": request.args.get("community", ""),
             "category": request.args.get("category", "")},
        teams=team_units(), towns=town_units(),
        categories=CATEGORIES,
        community_options=sorted(cnames), user=current_user())


@app.route("/export/ledger/settings", methods=["POST"])
@require_user
def ledger_settings():
    vals = []
    for key, _label, default in LEDGER_DEPS:
        val = (request.form.get(key) or "").strip() or default
        set_setting(key, val)
        vals.append(val)
    log_action("修改摸排台账表头", " / ".join(vals))
    flash("台账表头已保存", "ok")
    return redirect(url_for(
        "export_ledger",
        start=request.args.get("start", ""), end=request.args.get("end", ""),
        town=request.args.get("town", ""), team=request.args.get("team", ""),
        community=request.args.get("community", ""),
        category=request.args.get("category", ""),
        group=request.args.get("group", "")))


def _ledger_workbook(groups):
    """生成摸排台账工作簿：单表连续排版，页脚与原表一致；照片用 WPS 嵌入单元格（DISPIMG）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter
    from PIL import Image as PILImage

    deps = [get_setting(k, d) or d for k, _l, d in LEDGER_DEPS]
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # 字体仿原表：标题黑体28、表头黑体（序号/备注16，其余12）、正文宋体11
    title_font = Font(name="黑体", size=28)
    title_align = Alignment(horizontal="centerContinuous", vertical="center",
                            wrap_text=True)
    head_big = Font(name="黑体", size=16)
    head_small = Font(name="黑体", size=12)
    tno_font = Font(name="宋体", size=10.5)   # 表格序号行
    label_font = Font(name="宋体", size=24)   # 整改前/整改后标签

    headers = ["序号", "问题上报部门", "牵头部门", "配合部门",
               "问题描述", "整改举措", "整改时限", "整改情况", "备注"]
    widths = [7.375, 14.625, 21.125, 14.375, 20.75, 21.125, 14.625, 13.125, 9.125]
    def embed_photo(filepath):
        """直接取原图（不缩放、不垫白底），返回 (bytes, 宽, 高)。"""
        p = UPLOADS_DIR / filepath
        if not p.exists():
            return None
        try:
            with PILImage.open(p) as im:
                w, h = im.size
            return p.read_bytes(), w, h
        except Exception:
            return None

    wb = Workbook()
    ws = wb.active
    ws.title = "小区摸排台账"
    # 纸张：A4 横向 + 原表页边距
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_margins.left = ws.page_margins.right = 0.590277777777778
    ws.page_margins.top = ws.page_margins.bottom = 0.751388888888889
    ws.page_margins.header = ws.page_margins.footer = 0.298611111111111
    # 页脚与原表一致：居中「第 &P 页」
    ws.oddFooter.center.text = "第 &P 页"
    ws.oddFooter.center.size = 9
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.column_dimensions["J"].width = 9.64  # 原表右侧余量列
    if not groups:  # 无数据时也要有可见内容
        ws["A1"] = "当前范围内没有巡查记录"
        ws["A1"].font = title_font
    # 单表连续排版：下一个小区的表格接在上一个小区照片下面，不强制分页
    placements = []  # 嵌入单元格图片 {ref, disp, x, y, cx, cy, png}
    y_pt = 0.0       # 当前行顶距表顶的点数（算图片 y 偏移）
    r = 1
    for g in groups:
        # 标题行（黑体28 居中）
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        c = ws.cell(row=r, column=1, value=f"{g['community']}小区摸排情况")
        c.font = title_font
        c.alignment = title_align
        ws.row_dimensions[r].height = 35.25
        r += 1
        y_pt += 35.25
        # 表头（序号/备注黑体16，其余黑体12）
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=col, value=h)
            cell.font = head_big if col in (1, 9) else head_small
            cell.border = border
            cell.alignment = center
        ws.row_dimensions[r].height = 35
        r += 1
        y_pt += 35
        # 数据行：固定 9 行序号空间，不足补空行；超过 9 行顺延
        data_start = r
        for num, rec in g["rows"]:
            vals = [num, deps[0],
                    rec.get("lead_dept") or deps[1],
                    rec.get("assist_dept") or deps[2],
                    rec["description"] or rec["category"],
                    rec["result"] or "",
                    rec["deadline"] or "",
                    rec["_status_text"],
                    rec["_remark"] or ""]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=col, value=v)
                cell.border = border
                cell.font = Font(name="宋体", size=10 if col == 5 else 11)
                cell.alignment = center
            ws.row_dimensions[r].height = 45
            r += 1
            y_pt += 45
        while r < data_start + 9:  # 补足固定 9 行空间（带边框空行）
            for col in range(1, 10):
                ws.cell(row=r, column=col).border = border
            ws.row_dimensions[r].height = 45
            r += 1
            y_pt += 45
        # 每组照片一块：表格序号（共用分类序号）+ 整改前/整改后标签 + 照片
        for blk in g["blocks"]:
            tno = ws.cell(row=r, column=1,
                          value=f"{g['community']}表格序号{blk['num']}")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            tno.font = tno_font
            tno.alignment = center
            ws.row_dimensions[r].height = 45
            r += 1
            y_pt += 45
            lab = ws.cell(row=r, column=1, value="整改前")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            lab2 = ws.cell(row=r, column=6, value="整改后")
            ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=9)
            lab.font = lab2.font = label_font
            lab.alignment = lab2.alignment = center
            ws.row_dimensions[r].height = 45
            r += 1
            y_pt += 45
            for i in range(max(len(blk["before"]), len(blk["after"]))):
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
                ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=9)
                n = len(placements) + 1
                if i < len(blk["before"]):
                    res = embed_photo(blk["before"][i])
                    if res:
                        data, w, h = res
                        disp = f"ID_cg{n}"
                        placements.append({
                            "ref": f"A{r}", "disp": disp,
                            "x": 0, "y": int(round(y_pt * 12700)),
                            "cx": w * 9525, "cy": h * 9525,
                            "img": data,
                        })
                        ws.cell(row=r, column=1).value = f'=_xlfn.DISPIMG("{disp}",1)'
                        n += 1
                if i < len(blk["after"]):
                    res = embed_photo(blk["after"][i])
                    if res:
                        data, w, h = res
                        disp = f"ID_cg{n}"
                        placements.append({
                            "ref": f"F{r}", "disp": disp,
                            "x": 5962650, "y": int(round(y_pt * 12700)),
                            "cx": w * 9525, "cy": h * 9525,
                            "img": data,
                        })
                        ws.cell(row=r, column=6).value = f'=_xlfn.DISPIMG("{disp}",1)'
                ws.row_dimensions[r].height = 368.5
                r += 1
                y_pt += 368.5
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    if placements:
        buf = _inject_cellimages(buf, placements)
    return buf


def _inject_cellimages(buf, placements):
    """向 xlsx 注入 WPS「嵌入单元格」图片：cellimages.xml + DISPIMG 公式缓存值 + 媒体。"""
    import zipfile as _zip
    src = _zip.ZipFile(buf)
    out = io.BytesIO()
    with _zip.ZipFile(out, "w", _zip.ZIP_DEFLATED) as dst:
        for n in src.namelist():
            data = src.read(n)
            if n == "xl/worksheets/sheet1.xml":
                text = data.decode("utf-8")
                for pl in placements:
                    ref, disp = pl["ref"], pl["disp"]
                    # 公式单元格加 t="str" 与缓存值（同 WPS 原表写法）
                    m = re.search(r'<c r="%s"([^>]*)>' % re.escape(ref), text)
                    if m:
                        text = text.replace(
                            m.group(0),
                            '<c r="%s"%s t="str">' % (ref, m.group(1)), 1)
                    fstr = '_xlfn.DISPIMG("%s",1)' % disp
                    # openpyxl 写 <f>...</f><v></v> 或 <f>...</f></c>，两种情况都补上缓存值
                    text = text.replace(
                        "<f>%s</f><v></v>" % fstr,
                        '<f>%s</f><v>=DISPIMG("%s",1)</v>' % (fstr, disp),
                        1)
                    text = text.replace(
                        "<f>%s</f></c>" % fstr,
                        '<f>%s</f><v>=DISPIMG("%s",1)</v></c>' % (fstr, disp),
                        1)
                data = text.encode("utf-8")
            elif n == "xl/_rels/workbook.xml.rels":
                text = data.decode("utf-8")
                if "cellImage" not in text:
                    text = text.replace(
                        "</Relationships>",
                        '<Relationship Id="rIdCellImages" '
                        'Type="http://www.wps.cn/officeDocument/2020/cellImage" '
                        'Target="cellimages.xml"/></Relationships>')
                data = text.encode("utf-8")
            elif n == "[Content_Types].xml":
                text = data.decode("utf-8")
                if "cellimages.xml" not in text:
                    text = text.replace(
                        "</Types>",
                        '<Override PartName="/xl/cellimages.xml" '
                        'ContentType="application/vnd.wps-officedocument.cellimage+xml"/>'
                        "</Types>")
                if 'Extension="jpeg"' not in text:
                    text = text.replace(
                        "</Types>",
                        '<Default Extension="jpeg" ContentType="image/jpeg"/></Types>')
                data = text.encode("utf-8")
            dst.writestr(n, data)
        # cellimages.xml + 关系 + 媒体
        pics, rels = [], []
        for i, pl in enumerate(placements, 1):
            rels.append(
                '<Relationship Id="rId%d" Type="http://schemas.openxmlformats.org'
                '/officeDocument/2006/relationships/image" Target="media/image%d.jpeg"/>'
                % (i, i))
            pics.append(
                "<etc:cellImage><xdr:pic>"
                + '<xdr:nvPicPr><xdr:cNvPr id="%d" name="%s"/>'
                % (1000 + i, pl["disp"])
                + '<xdr:cNvPicPr><a:picLocks noChangeAspect="1"/></xdr:cNvPicPr>'
                + "</xdr:nvPicPr>"
                + '<xdr:blipFill><a:blip r:embed="rId%d"/>' % i
                + "<a:stretch><a:fillRect/></a:stretch></xdr:blipFill>"
                + '<xdr:spPr><a:xfrm><a:off x="%d" y="%d"/>'
                % (pl["x"], pl["y"])
                + '<a:ext cx="%d" cy="%d"/></a:xfrm>' % (pl["cx"], pl["cy"])
                + '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
                + '<a:noFill/><a:ln w="9525"><a:noFill/></a:ln></xdr:spPr>'
                + "</xdr:pic></etc:cellImage>")
        ci = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            "<etc:cellImages "
            'xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
            'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
            'xmlns:etc="http://www.wps.cn/officeDocument/2017/etCustomData">'
            + "".join(pics) + "</etc:cellImages>")
        dst.writestr("xl/cellimages.xml", ci.encode("utf-8"))
        rels_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org'
            '/package/2006/relationships">' + "".join(rels)
            + "</Relationships>")
        dst.writestr("xl/_rels/cellimages.xml.rels", rels_xml.encode("utf-8"))
        for i, pl in enumerate(placements, 1):
            dst.writestr("xl/media/image%d.jpeg" % i, pl["img"])
    out.seek(0)
    return out

def _ledger_scope_label(unit=None):
    """按已选条件拼名称片段：时间 至 时间 + 乡镇/中队 + 小区。"""
    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()
    town = (request.args.get("town") or "").strip()
    team = (request.args.get("team") or "").strip()
    community = (request.args.get("community") or "").strip()
    parts = []
    if start and end:
        parts.append(f"{start}至{end}")
    elif start or end:
        parts.append(start or end)
    if unit:
        parts.append(unit)
    elif team:
        parts.append(team)
    elif town:
        parts.append(town)
    if community:
        parts.append(community)
    return " ".join(parts) if parts else "全部"


@app.route("/export/ledger.xlsx")
@require_user
def export_ledger_xlsx():
    import zipfile as _zip
    group = request.args.get("group", "")
    community = (request.args.get("community") or "").strip()
    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()

    def date_part():
        if start and end:
            return f"{start}至{end}"
        return start or end or ""

    if group in ("team", "town"):
        # 按中队/乡镇分组：每个单位一个文件夹，文件夹内放该单位的台账 Excel
        units = team_units() if group == "team" else town_units()
        town = (request.args.get("town") or "").strip()
        team = (request.args.get("team") or "").strip()
        if group == "team" and town:
            units = [t for t in units if t.startswith(town)]
        if group == "team" and team:
            units = [t for t in units if t == team]
        if group == "town" and team:
            units = [t for t in units if team.startswith(t)]
        zbuf = io.BytesIO()
        total = 0
        with _zip.ZipFile(zbuf, "w", _zip.ZIP_DEFLATED) as z:
            for unit in units:
                where, params = scope_where()
                cond = "team = ?" if group == "team" else "team LIKE ?"
                val = unit if group == "team" else unit + "%"
                where = (where + " AND " if where else "") + cond
                params.append(val)
                if start:
                    where = (where + " AND " if where else "") + "created_at >= ?"
                    params.append(start + " 00:00")
                if end:
                    where = (where + " AND " if where else "") + "created_at <= ?"
                    params.append(end + " 23:59")
                if community:
                    where = (where + " AND " if where else "") + "community = ?"
                    params.append(community)
                category = (request.args.get("category") or "").strip()
                if category:
                    where = (where + " AND " if where else "") + "category = ?"
                    params.append(category)
                groups = _ledger_groups_from(where, params, unit=unit,
                                             unit_town=(group == "town"))
                if not groups:
                    continue
                d = date_part()
                z.writestr(
                    f"{unit}/{(d + ' ') if d else ''}{unit}小区摸排台账.xlsx",
                    _ledger_workbook(groups).getvalue())
                total += len(groups)
        if total == 0:
            zbuf = io.BytesIO()
            with _zip.ZipFile(zbuf, "w", _zip.ZIP_DEFLATED) as z:
                z.writestr("无数据.txt", "当前筛选范围内没有巡查记录")
        zbuf.seek(0)
        label = "按中队" if group == "team" else "按乡镇"
        fname = f"{_ledger_scope_label()}小区摸排台账_{label}.zip".replace("/", "-")
        log_action("导出摸排台账", f"{label} · {fname} · {total} 个小区")
        return send_file(zbuf, as_attachment=True, download_name=fname,
                         mimetype="application/zip")

    groups, _, _ = _ledger_groups()
    buf = _ledger_workbook(groups)
    fname = f"{_ledger_scope_label()}小区摸排台账.xlsx".replace("/", "-")
    log_action("导出摸排台账", f"{fname} · {len(groups)} 个小区")
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# 单条记录打印：按台账格式导出仅此一条（打印用）
@app.route("/record/<int:rid>/ledger.xlsx")
@require_user
def record_ledger(rid):
    r = get_db().execute("SELECT * FROM records WHERE id=?", (rid,)).fetchone()
    if not r:
        abort(404)
    if not can_view_record(r):
        abort(403)
    r = dict(r)
    r["_status_text"] = "已整改" if r["status"] == "closed" else "未整改"
    r["_remark"] = ""
    imgs = get_db().execute(
        "SELECT * FROM images WHERE record_id=? ORDER BY id", (rid,)
    ).fetchall()
    before = [im["filepath"] for im in imgs if im["type"] == "before"]
    after = [im["filepath"] for im in imgs if im["type"] == "after"]
    blocks = []
    if before or after:
        blocks.append({"num": 1, "before": before, "after": after})
    groups = [{
        "community": r["community"] or "未填小区",
        "rows": [(1, r)],
        "pad": 8,
        "blocks": blocks,
    }]
    buf = _ledger_workbook(groups)
    fname = f"{(r['community'] or '未填小区')}_{r['category']}_记录{rid}.xlsx".replace("/", "-")
    log_action("打印单条记录", f"记录#{rid} {r['community'] or '未填小区'} · {r['category']}")
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------- 数据一键备份（主管理员） ----------
@app.route("/admin/backup")
@require_user
def admin_backup():
    u = current_user()
    if not u or u["role"] != "super":
        abort(403)
    import zipfile as _zip
    buf = io.BytesIO()
    with _zip.ZipFile(buf, "w", _zip.ZIP_DEFLATED) as z:
        db_path = DATA_DIR / "records.db"
        if db_path.exists():
            z.write(db_path, "data/records.db")
        if UPLOADS_DIR.exists():
            for f in UPLOADS_DIR.rglob("*"):
                if f.is_file():
                    z.write(f, "uploads/" + str(f.relative_to(UPLOADS_DIR)))
    buf.seek(0)
    fname = f"chengguan_backup_{datetime.now().strftime('%Y%m%d_%H%M')}.zip"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/zip")


# ---------- 辅助接口 ----------
@app.route("/api/communities")
@require_user
def api_communities():
    q = request.args.get("q", "").strip()
    return jsonify(community_autocomplete(q))


@app.route("/uploads/<path:filepath>")
def uploads(filepath):
    return send_from_directory(UPLOADS_DIR, filepath)


@app.route("/manifest.json")
def manifest():
    return send_from_directory(BASE_DIR, "manifest.json",
                               mimetype="application/manifest+json")


@app.route("/sw.js")
def sw():
    return send_from_directory(BASE_DIR, "sw.js",
                               mimetype="application/javascript")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
